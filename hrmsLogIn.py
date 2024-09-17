from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import random
import string
import time
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Path to your chromedriver
chromedriver_path = 'C:/Users/roni58625/.cache/selenium/chromedriver/win64/91.0.4472.101/chromedriver.exe'

# Function to get the last updated password from the Excel file
def get_last_password(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Find the last non-empty row in the first column
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=1).value is not None:
            last_password = sheet.cell(row=row, column=1).value
            return last_password

    return None  # In case the Excel sheet is empty

# Function to generate a new password
def generate_password():
    # Password criteria: 6 characters, at least one uppercase, one lowercase, one digit, and one special character '@'
    while True:
        password = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits + '@', k=6))
        if (any(c.islower() for c in password) and any(c.isupper() for c in password)
                and any(c.isdigit() for c in password) and '@' in password):
            return password

# Function to update the Excel file with the new password
def update_excel(new_password, excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Find the first truly empty row (not just max_row)
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value is None:
            next_row = row
            break
    else:
        next_row = sheet.max_row + 1  # If no empty row found, add to the next row

    # Save the new password in the first column, and timestamp in the second column
    sheet.cell(row=next_row, column=1, value=new_password)
    sheet.cell(row=next_row, column=2, value=time.strftime('%Y-%m-%d %H:%M:%S'))

    # Save changes to the workbook
    workbook.save(excel_file)
    print(f"Password updated in row {next_row}.")
    
    return next_row  # Return the row number of the newly added password

# Function to send an email notification
def send_email(new_password, old_password, update_time):
    try:
        # Email configuration
        sender_email = "rahatuzzaman2000@gmail.com"
        recipient_email = "die.r&i_designer7@waltonbd.com"
        subject = "Password Updated"
        
        # Create the email body
        body = (
            f"The HRMS password has been updated to: {new_password}\n\n"
            f"Update Time: {update_time}\n\n"
            f"Previous password: {old_password}\n"
        )
        
        # Create a MIMEText object
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = recipient_email
        message["Subject"] = subject
        
        # Attach the body to the email
        message.attach(MIMEText(body, "plain"))
        
        # Connect to the SMTP server and send the email
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()  # Upgrade the connection to secure
            server.login(sender_email, "gqck wxrq wmaq fxck")  # Use your App Password here
            server.sendmail(sender_email, recipient_email, message.as_string())
            print("Email sent successfully.")

    except Exception as e:
        print(f"Error sending email: {e}")

# Set up the Chrome WebDriver
service = Service(executable_path=chromedriver_path)
driver = webdriver.Chrome(service=service)

# Open the HRMS login page
driver.get("https://hrms.waltonbd.com/HRMS/")

# Wait for the page to load completely
time.sleep(5)

# Login process
wait = WebDriverWait(driver, 10)

# Enter user ID
try:
    username_field = wait.until(EC.presence_of_element_located((By.ID, 'username')))
    username_field.send_keys("58625")
    print("User ID entered.")
except Exception as e:
    print(f"Error locating User ID field: {e}")

# Get the last updated password from the Excel file
excel_file = os.path.join("C:\\Users\\roni58625\\Desktop\\Python Script\\dist", "hrms.xlsx")
old_password = get_last_password(excel_file)

# Enter password (from Excel)
try:
    password_field = wait.until(EC.presence_of_element_located((By.ID, 'password')))
    password_field.send_keys(old_password)  # Use the password from Excel
    print("Password entered from Excel.")
except Exception as e:
    print(f"Error locating Password field: {e}")

# Click the login button
try:
    login_button = wait.until(EC.element_to_be_clickable((By.NAME, "_action_authenticate")))
    login_button.click()
    print("Login button clicked.")
except Exception as e:
    print(f"Error locating Login button: {e}")

# Wait for the page to load after login
time.sleep(3)

# Navigate to "Self Service" -> "Change Emp Password"
try:
    self_service_dropdown = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Self Service")))
    self_service_dropdown.click()
    print("Self Service dropdown clicked.")
except Exception as e:
    print(f"Error locating Self Service dropdown: {e}")

try:
    change_emp_password_option = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Change Emp Password")))
    change_emp_password_option.click()
    print("Change Emp Password option clicked.")
except Exception as e:
    print(f"Error locating Change Emp Password option: {e}")

# Wait for the password change page to load
time.sleep(3)

# Enter old password
try:
    old_password_field = wait.until(EC.presence_of_element_located((By.ID, 'oldPassword')))
    old_password_field.send_keys(old_password)
    print("Old password entered.")
except Exception as e:
    print(f"Error locating old password field: {e}")

# Generate and enter new password
new_password = generate_password()
try:
    new_password_field = wait.until(EC.presence_of_element_located((By.ID, 'newPassword')))
    new_password_field.send_keys(new_password)
    print("New password entered.")
except Exception as e:
    print(f"Error locating new password field: {e}")

# Retype new password
try:
    re_new_password_field = wait.until(EC.presence_of_element_located((By.ID, 'reNewPassword')))
    re_new_password_field.send_keys(new_password)
    print("Retyped new password entered.")
except Exception as e:
    print(f"Error locating confirmation password field: {e}")

# Click the Update button to submit the new password
try:
    update_button = wait.until(EC.element_to_be_clickable((By.ID, 'create')))
    update_button.click()
    print("Password change submitted.")
    
    # Handle confirmation pop-up
    alert = WebDriverWait(driver, 10).until(EC.alert_is_present())
    alert.accept()
    print("Confirmation pop-up accepted.")
    
except Exception as e:
    print(f"Error clicking the Update button or handling the pop-up: {e}")

# Update the Excel file with the new password and timestamp
try:
    row_num = update_excel(new_password, excel_file)
    print("Excel file updated with new password.")
except Exception as e:
    print(f"Error updating Excel file: {e}")

# Send an email notification
try:
    update_time = time.strftime('%Y-%m-%d %H:%M:%S')
    send_email(new_password, old_password, update_time)
except Exception as e:
    print(f"Error sending email: {e}")

# Close the browser after the process
time.sleep(5)  # Wait to visually confirm the change before closing
driver.quit()
